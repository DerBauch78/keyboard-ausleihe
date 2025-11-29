"""Excel-Export Funktionen für Keyboard-Ausleihe"""
import io
import json
import zipfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def style_header(ws, row=1, columns=None):
    """Header-Zeile formatieren"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    
    for col in range(1, (columns or ws.max_column) + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_column_width(ws):
    """Spaltenbreiten automatisch anpassen"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def export_json_backup(school_year):
    """Komplettes Backup als JSON (für Reimport)"""
    from app.models import SchoolClass, Student, Keyboard, Loan, User
    
    data = {
        "export_version": "2.0",
        "exported_at": datetime.now().isoformat(),
        "school_year": {
            "name": school_year.name,
            "start_date": school_year.start_date.isoformat(),
            "end_date": school_year.end_date.isoformat(),
            "is_active": school_year.is_active
        },
        "keyboards": [],
        "classes": [],
        "loans": []
    }
    
    # Keyboards
    for kb in Keyboard.query.order_by(Keyboard.internal_number).all():
        data["keyboards"].append({
            "inventory_number": kb.inventory_number,
            "internal_number": kb.internal_number,
            "condition": kb.condition,
            "status": kb.status,
            "notes": kb.notes
        })
    
    # Klassen und Schüler
    for cls in school_year.classes.order_by(SchoolClass.name).all():
        class_data = {
            "name": cls.name,
            "grade": cls.grade,
            "class_teacher": cls.class_teacher,
            "music_teacher": cls.music_teacher,
            "students": []
        }
        
        for student in cls.students.order_by(Student.last_name, Student.first_name).all():
            class_data["students"].append({
                "last_name": student.last_name,
                "first_name": student.first_name,
                "notes": student.notes,
                "participates_in_loan": student.participates_in_loan,
                "fee_prepaid": student.fee_prepaid
            })
        
        data["classes"].append(class_data)
    
    # Aktive Ausleihen
    for loan in Loan.query.filter(Loan.returned_at == None).all():
        data["loans"].append({
            "student_class": loan.student.school_class.name,
            "student_last_name": loan.student.last_name,
            "student_first_name": loan.student.first_name,
            "keyboard_inventory_number": loan.keyboard.inventory_number,
            "loaned_at": loan.loaned_at.isoformat() if loan.loaned_at else None,
            "fee_paid": loan.fee_paid,
            "fee_amount": loan.fee_amount
        })
    
    return json.dumps(data, ensure_ascii=False, indent=2)


def export_full_backup_zip(school_year):
    """Komplettes Backup als ZIP mit Excel + JSON"""
    date_str = datetime.now().strftime('%Y%m%d')
    year_str = school_year.name.replace('/', '-')
    
    # Excel erstellen
    excel_output = export_full_backup(school_year)
    
    # JSON erstellen
    json_content = export_json_backup(school_year)
    
    # ZIP erstellen
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Excel hinzufügen
        zip_file.writestr(
            f"keyboard_backup_{year_str}_{date_str}.xlsx",
            excel_output.getvalue()
        )
        # JSON hinzufügen
        zip_file.writestr(
            f"keyboard_backup_{year_str}_{date_str}.json",
            json_content.encode('utf-8')
        )
    
    zip_buffer.seek(0)
    return zip_buffer


def export_full_backup(school_year):
    """Komplettes Backup eines Schuljahres als Excel"""
    from app.models import SchoolClass, Student, Keyboard, Loan
    
    wb = Workbook()
    
    # === Sheet 1: Übersicht ===
    ws_overview = wb.active
    ws_overview.title = "Übersicht"
    
    ws_overview['A1'] = "Keyboard-Ausleihe Backup"
    ws_overview['A1'].font = Font(bold=True, size=16)
    ws_overview['A2'] = f"Schuljahr: {school_year.name}"
    ws_overview['A3'] = f"Exportiert: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    
    # Statistiken
    ws_overview['A5'] = "Statistiken"
    ws_overview['A5'].font = Font(bold=True)
    
    total_keyboards = Keyboard.query.count()
    active_loans = Loan.query.filter(Loan.returned_at == None).count()
    paid_loans = Loan.query.filter(Loan.returned_at == None, Loan.fee_paid == True).count()
    
    stats = [
        ("Keyboards gesamt", total_keyboards),
        ("Aktive Ausleihen", active_loans),
        ("Gebühren bezahlt", paid_loans),
        ("Gebühren offen", active_loans - paid_loans),
        ("Einnahmen (bezahlt)", f"{paid_loans * 10}€"),
        ("Einnahmen (offen)", f"{(active_loans - paid_loans) * 10}€"),
    ]
    
    for i, (label, value) in enumerate(stats, start=6):
        ws_overview[f'A{i}'] = label
        ws_overview[f'B{i}'] = value
    
    # === Sheet 2: Alle Ausleihen (aktiv) ===
    ws_loans = wb.create_sheet("Aktive Ausleihen")
    
    headers = ["Klasse", "Nachname", "Vorname", "Keyboard", "Ausgeliehen am", "Gebühr bezahlt", "Anmerkungen"]
    ws_loans.append(headers)
    style_header(ws_loans)
    
    active_loans_data = Loan.query.filter(Loan.returned_at == None).join(Student).join(SchoolClass).order_by(
        SchoolClass.name, Student.last_name, Student.first_name
    ).all()
    
    for loan in active_loans_data:
        ws_loans.append([
            loan.student.school_class.name,
            loan.student.last_name,
            loan.student.first_name,
            loan.keyboard.inventory_number,
            loan.loaned_at.strftime('%d.%m.%Y'),
            "Ja" if loan.fee_paid else "Nein",
            loan.student.notes or ""
        ])
    
    auto_column_width(ws_loans)
    
    # === Sheet 3: Klassen mit Schülern ===
    for cls in school_year.classes.order_by(SchoolClass.name).all():
        ws_class = wb.create_sheet(f"Klasse {cls.name}")
        
        # Klasseninfo
        ws_class['A1'] = f"Klasse {cls.name}"
        ws_class['A1'].font = Font(bold=True, size=14)
        if cls.class_teacher:
            ws_class['A2'] = f"Klassenlehrer: {cls.class_teacher}"
        
        # Schülerliste
        headers = ["Nr.", "Nachname", "Vorname", "Keyboard", "Gebühr", "Anmerkungen"]
        ws_class.append([])  # Leerzeile
        ws_class.append(headers)
        style_header(ws_class, row=ws_class.max_row)
        
        students = cls.students.order_by(Student.last_name, Student.first_name).all()
        for i, student in enumerate(students, start=1):
            loan = student.current_loan
            ws_class.append([
                i,
                student.last_name,
                student.first_name,
                loan.keyboard.inventory_number if loan else "",
                "Bezahlt" if (loan and loan.fee_paid) else ("Offen" if loan else ""),
                student.notes or ""
            ])
        
        auto_column_width(ws_class)
    
    # === Sheet 4: Keyboard-Inventar ===
    ws_keyboards = wb.create_sheet("Keyboard-Inventar")
    
    headers = ["Nr.", "Inventarnummer", "Status", "Zustand", "Aktueller Ausleiher", "Klasse", "Notizen"]
    ws_keyboards.append(headers)
    style_header(ws_keyboards)
    
    keyboards = Keyboard.query.order_by(Keyboard.internal_number, Keyboard.inventory_number).all()
    for kb in keyboards:
        loan = kb.current_loan
        ws_keyboards.append([
            kb.internal_number or "",
            kb.inventory_number,
            dict(Keyboard.STATUS_CHOICES).get(kb.status, kb.status),
            dict(Keyboard.CONDITION_CHOICES).get(kb.condition, kb.condition),
            loan.student.full_name if loan else "",
            loan.student.school_class.name if loan else "",
            kb.notes or ""
        ])
    
    auto_column_width(ws_keyboards)
    
    # === Sheet 5: Rückgaben (Historie) ===
    ws_returned = wb.create_sheet("Rückgaben")
    
    headers = ["Klasse", "Nachname", "Vorname", "Keyboard", "Ausgeliehen", "Zurückgegeben", "Zustand", "Bemerkung"]
    ws_returned.append(headers)
    style_header(ws_returned)
    
    returned_loans = Loan.query.filter(Loan.returned_at != None).join(Student).join(SchoolClass).filter(
        SchoolClass.school_year_id == school_year.id
    ).order_by(Loan.returned_at.desc()).all()
    
    for loan in returned_loans:
        ws_returned.append([
            loan.student.school_class.name,
            loan.student.last_name,
            loan.student.first_name,
            loan.keyboard.inventory_number,
            loan.loaned_at.strftime('%d.%m.%Y'),
            loan.returned_at.strftime('%d.%m.%Y') if loan.returned_at else "",
            dict(Keyboard.CONDITION_CHOICES).get(loan.return_condition, "") if loan.return_condition else "",
            loan.return_notes or ""
        ])
    
    auto_column_width(ws_returned)
    
    # In BytesIO speichern
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_class_list(school_class):
    """Einzelne Klassenliste als Excel"""
    from app.models import Student
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Klasse {school_class.name}"
    
    # Header
    ws['A1'] = f"Klassenliste {school_class.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Schuljahr: {school_class.school_year.name}"
    if school_class.class_teacher:
        ws['A3'] = f"Klassenlehrer: {school_class.class_teacher}"
    
    # Tabelle
    headers = ["Nr.", "Nachname", "Vorname", "Keyboard", "Gebühr bezahlt", "Anmerkungen"]
    ws.append([])
    ws.append(headers)
    style_header(ws, row=ws.max_row)
    
    students = school_class.students.order_by(Student.last_name, Student.first_name).all()
    for i, student in enumerate(students, start=1):
        loan = student.current_loan
        ws.append([
            i,
            student.last_name,
            student.first_name,
            loan.keyboard.inventory_number if loan else "",
            "Ja" if (loan and loan.fee_paid) else ("Nein" if loan else ""),
            student.notes or ""
        ])
    
    auto_column_width(ws)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_payment_list(school_year):
    """Gebühren-Übersicht als Excel (für Buchhaltung)"""
    from app.models import SchoolClass, Student, Loan
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Gebühren"
    
    ws['A1'] = "Gebühren-Übersicht Keyboard-Ausleihe"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Schuljahr: {school_year.name}"
    ws['A3'] = f"Stand: {datetime.now().strftime('%d.%m.%Y')}"
    
    headers = ["Klasse", "Nachname", "Vorname", "Betrag", "Status", "Keyboard"]
    ws.append([])
    ws.append(headers)
    style_header(ws, row=ws.max_row)
    
    # Rot für offen, grün für bezahlt
    red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    active_loans = Loan.query.filter(Loan.returned_at == None).join(Student).join(SchoolClass).filter(
        SchoolClass.school_year_id == school_year.id
    ).order_by(SchoolClass.name, Student.last_name).all()
    
    total_paid = 0
    total_open = 0
    
    for loan in active_loans:
        row = [
            loan.student.school_class.name,
            loan.student.last_name,
            loan.student.first_name,
            "10,00 €",
            "Bezahlt" if loan.fee_paid else "Offen",
            loan.keyboard.inventory_number
        ]
        ws.append(row)
        
        # Zeile einfärben
        current_row = ws.max_row
        fill = green_fill if loan.fee_paid else red_fill
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).fill = fill
        
        if loan.fee_paid:
            total_paid += 10
        else:
            total_open += 10
    
    # Summen
    ws.append([])
    ws.append(["", "", "Summe bezahlt:", f"{total_paid},00 €"])
    ws.append(["", "", "Summe offen:", f"{total_open},00 €"])
    ws.append(["", "", "Gesamt:", f"{total_paid + total_open},00 €"])
    
    for row in range(ws.max_row - 2, ws.max_row + 1):
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4).font = Font(bold=True)
    
    auto_column_width(ws)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
